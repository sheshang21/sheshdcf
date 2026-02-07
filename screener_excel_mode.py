"""
SCREENER EXCEL MODE - Auxiliary Module for SheshUltimate DCF Valuation System
==============================================================================

This module handles Screener.in template format Excel files with:
- Balance Sheet (rows with item names in column A, years in row 2)
- Profit and Loss Account (rows with item names in column A, years in row 2)

Features:
- Full DCF Valuation
- DDM (Dividend Discount Model) 
- RIM (Residual Income Model)
- Peer Comparison Dashboard
- Comparative Valuation
- Excel Download Functionality
- All unlisted mode features adapted for Screener template format

Author: SheshUltimate Team
Version: 1.0.0
"""

import pandas as pd
import numpy as np
import streamlit as st
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
from io import BytesIO


# ================================
# SCREENER EXCEL PARSING FUNCTIONS
# ================================

def parse_screener_excel_to_dataframes(excel_file):
    """
    Parse Screener.in template Excel file
    
    Expected Structure:
    - Sheet 1: 'Balance Sheet' with years in Row 2, items in Column A
    - Sheet 2: 'Profit and Loss Account' with years in Row 2, items in Column A
    
    Returns:
        df_bs, df_pl: DataFrames with Item column and year columns
    """
    try:
        # Read both sheets - note exact sheet names from template
        df_bs = pd.read_excel(excel_file, sheet_name='Balance Sheet', header=None)
        df_pl = pd.read_excel(excel_file, sheet_name='Profit and Loss Account', header=None)
        
        # Process Balance Sheet
        # Row 0: Title "BALANCE SHEET"
        # Row 1: "Report Date" + actual dates
        # Row 2+: Item names in column 0, values in columns 1+
        
        # Get years from Row 1 (index 1)
        bs_dates = df_bs.iloc[1, 1:].values  # Skip column 0 ("Report Date")
        
        # Convert dates to year strings
        bs_years = []
        for date_val in bs_dates:
            if pd.notna(date_val):
                if isinstance(date_val, datetime):
                    bs_years.append(f'_{date_val.year}')
                elif isinstance(date_val, str):
                    # Try to extract year from string
                    try:
                        year = pd.to_datetime(date_val).year
                        bs_years.append(f'_{year}')
                    except:
                        bs_years.append(f'_col{len(bs_years)+1}')
                else:
                    bs_years.append(f'_col{len(bs_years)+1}')
            else:
                bs_years.append(f'_col{len(bs_years)+1}')
        
        # Same for P&L
        pl_dates = df_pl.iloc[1, 1:].values
        pl_years = []
        for date_val in pl_dates:
            if pd.notna(date_val):
                if isinstance(date_val, datetime):
                    pl_years.append(f'_{date_val.year}')
                elif isinstance(date_val, str):
                    try:
                        year = pd.to_datetime(date_val).year
                        pl_years.append(f'_{year}')
                    except:
                        pl_years.append(f'_col{len(pl_years)+1}')
                else:
                    pl_years.append(f'_col{len(pl_years)+1}')
            else:
                pl_years.append(f'_col{len(pl_years)+1}')
        
        # Create column names: 'Item' + year columns
        bs_columns = ['Item'] + bs_years
        pl_columns = ['Item'] + pl_years
        
        # Extract data starting from Row 2 (index 2)
        df_bs_data = df_bs.iloc[2:].copy()
        df_pl_data = df_pl.iloc[2:].copy()
        
        # Limit to actual columns
        df_bs_data = df_bs_data.iloc[:, :len(bs_columns)]
        df_pl_data = df_pl_data.iloc[:, :len(pl_columns)]
        
        # Set column names
        df_bs_data.columns = bs_columns
        df_pl_data.columns = pl_columns
        
        # Reset index
        df_bs_data = df_bs_data.reset_index(drop=True)
        df_pl_data = df_pl_data.reset_index(drop=True)
        
        # Convert year columns to numeric
        for col in bs_years:
            df_bs_data[col] = pd.to_numeric(df_bs_data[col], errors='coerce').fillna(0.0)
        
        for col in pl_years:
            df_pl_data[col] = pd.to_numeric(df_pl_data[col], errors='coerce').fillna(0.0)
        
        # Remove rows where Item is NaN or empty
        df_bs_data = df_bs_data[df_bs_data['Item'].notna() & (df_bs_data['Item'] != '')]
        df_pl_data = df_pl_data[df_pl_data['Item'].notna() & (df_pl_data['Item'] != '')]
        
        # Clean Item column - remove whitespace
        df_bs_data['Item'] = df_bs_data['Item'].astype(str).str.strip()
        df_pl_data['Item'] = df_pl_data['Item'].astype(str).str.strip()
        
        return df_bs_data, df_pl_data
        
    except Exception as e:
        st.error(f"❌ Error parsing Screener Excel: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None, None


def get_value_from_screener_df(df, item_name, year_col):
    """
    Extract value from Screener DataFrame by item name (case-insensitive partial match)
    
    Args:
        df: DataFrame with 'Item' column and year columns
        item_name: Item to search for (partial match, case-insensitive)
        year_col: Year column name (e.g., '_2023')
    
    Returns:
        float: Value or 0.0 if not found
    """
    if df is None or df.empty:
        return 0.0
    
    item_name_lower = item_name.lower()
    mask = df['Item'].str.lower().str.contains(item_name_lower, na=False, regex=False)
    matching = df[mask]
    
    if not matching.empty and year_col in matching.columns:
        return float(matching.iloc[0][year_col])
    return 0.0


def detect_screener_year_columns(df):
    """
    Detect year columns dynamically (columns starting with _)
    
    Returns:
        list: Sorted list of year column names
    """
    if df is None or df.empty:
        return []
    
    year_cols = [col for col in df.columns if col.startswith('_') and col != 'Item']
    # Sort by numeric value after underscore
    year_cols.sort(key=lambda x: int(x[1:]) if x[1:].isdigit() else 0)
    return year_cols


# ================================
# SCREENER FINANCIAL EXTRACTION
# ================================

def extract_screener_financials(df_bs, df_pl, year_cols):
    """
    Extract financial metrics from Screener Excel DataFrames
    
    Maps Screener template items to financial metrics:
    
    Balance Sheet Items:
    - Equity Share Capital
    - Reserves
    - Borrowings (Debt)
    - Net Block (Fixed Assets)
    - Investments
    - Receivables
    - Inventory
    - Cash & Bank
    
    P&L Items:
    - Sales (Revenue)
    - Raw Material Cost, Change in Inventory, Power and Fuel, Other Mfr. Exp (COGS components)
    - Employee Cost, Selling and admin (OpEx)
    - Other Income, Other Expenses
    - Depreciation
    - Interest
    - Tax
    - Net profit
    - Dividend Amount
    
    Returns:
        dict: Financial metrics with 3 years of historical data (NEWEST FIRST)
    """
    num_years = min(3, len(year_cols))
    last_years = year_cols[-num_years:]
    
    # REVERSE to match unlisted mode - NEWEST FIRST [0], OLDEST LAST [-1]
    last_years = list(reversed(last_years))
    
    financials = {
        'years': last_years,
        'revenue': [],
        'cogs': [],
        'opex': [],
        'ebitda': [],
        'depreciation': [],
        'ebit': [],
        'interest': [],
        'interest_income': [],  # For business classification
        'tax': [],
        'nopat': [],
        'net_profit': [],  # Actual reported net profit
        'dividends': [],  # Dividend amounts
        'fixed_assets': [],
        'inventory': [],
        'receivables': [],
        'payables': [],
        'cash': [],
        'equity': [],
        'st_debt': [],
        'lt_debt': [],
    }
    
    for year_col in last_years:
        # ===== INCOME STATEMENT =====
        
        # Revenue (Sales)
        revenue = get_value_from_screener_df(df_pl, 'Sales', year_col)
        
        # COGS Components
        raw_material = get_value_from_screener_df(df_pl, 'Raw Material Cost', year_col)
        change_inventory = get_value_from_screener_df(df_pl, 'Change in Inventory', year_col)
        power_fuel = get_value_from_screener_df(df_pl, 'Power and Fuel', year_col)
        other_mfr = get_value_from_screener_df(df_pl, 'Other Mfr', year_col)
        
        # Note: Change in inventory can be negative (increase) or positive (decrease)
        # COGS = Raw Material + Change in Inventory + Power & Fuel + Other Manufacturing
        cogs = raw_material + change_inventory + power_fuel + other_mfr
        
        # Operating Expenses
        employee_cost = get_value_from_screener_df(df_pl, 'Employee Cost', year_col)
        selling_admin = get_value_from_screener_df(df_pl, 'Selling and admin', year_col)
        
        opex = employee_cost + selling_admin
        
        # Other Income and Expenses
        other_income = get_value_from_screener_df(df_pl, 'Other Income', year_col)
        other_expenses = get_value_from_screener_df(df_pl, 'Other Expenses', year_col)
        
        # Depreciation
        depreciation = get_value_from_screener_df(df_pl, 'Depreciation', year_col)
        
        # Interest
        interest = get_value_from_screener_df(df_pl, 'Interest', year_col)
        
        # Note: Screener template doesn't have separate interest income
        # If other_income is primarily interest income, it could be used for classification
        interest_income = 0.0  # Default to 0 unless template provides it
        
        # Tax
        tax = get_value_from_screener_df(df_pl, 'Tax', year_col)
        
        # Net Profit (reported)
        net_profit = get_value_from_screener_df(df_pl, 'Net profit', year_col)
        
        # Dividends
        dividend_amount = get_value_from_screener_df(df_pl, 'Dividend Amount', year_col)
        
        # Calculate EBITDA and EBIT
        # EBITDA = Revenue - COGS - OpEx + Other Income - Other Expenses
        ebitda = revenue - cogs - opex + other_income - other_expenses
        
        # EBIT = EBITDA - Depreciation
        ebit = ebitda - depreciation
        
        # NOPAT = EBIT * (1 - tax_rate)
        # Estimate tax rate from reported figures
        pbt = get_value_from_screener_df(df_pl, 'Profit before tax', year_col)
        if pbt > 0 and tax > 0:
            effective_tax_rate = tax / pbt
        else:
            effective_tax_rate = 0.25  # Default 25%
        
        nopat = ebit * (1 - effective_tax_rate)
        
        # Store P&L metrics
        financials['revenue'].append(revenue)
        financials['cogs'].append(cogs)
        financials['opex'].append(opex)
        financials['ebitda'].append(ebitda)
        financials['depreciation'].append(depreciation)
        financials['ebit'].append(ebit)
        financials['interest'].append(interest)
        financials['interest_income'].append(interest_income)
        financials['tax'].append(tax)
        financials['nopat'].append(nopat)
        financials['net_profit'].append(net_profit)
        financials['dividends'].append(dividend_amount)
        
        # ===== BALANCE SHEET =====
        
        # Fixed Assets (Net Block + Capital Work in Progress)
        net_block = get_value_from_screener_df(df_bs, 'Net Block', year_col)
        cwip = get_value_from_screener_df(df_bs, 'Capital Work in Progress', year_col)
        fixed_assets = net_block + cwip
        
        # Current Assets
        inventory = get_value_from_screener_df(df_bs, 'Inventory', year_col)
        receivables = get_value_from_screener_df(df_bs, 'Receivables', year_col)
        cash = get_value_from_screener_df(df_bs, 'Cash & Bank', year_col)
        
        # Equity (Equity Share Capital + Reserves)
        share_capital = get_value_from_screener_df(df_bs, 'Equity Share Capital', year_col)
        reserves = get_value_from_screener_df(df_bs, 'Reserves', year_col)
        equity = share_capital + reserves
        
        # Debt (Borrowings)
        borrowings = get_value_from_screener_df(df_bs, 'Borrowings', year_col)
        # Screener template has only total borrowings, split as 60% LT, 40% ST
        lt_debt = borrowings * 0.6
        st_debt = borrowings * 0.4
        
        # Payables (from Other Liabilities - rough estimate)
        other_liabilities = get_value_from_screener_df(df_bs, 'Other Liabilities', year_col)
        # Assume 70% of other liabilities are trade payables
        payables = other_liabilities * 0.7
        
        # Store Balance Sheet metrics
        financials['fixed_assets'].append(fixed_assets)
        financials['inventory'].append(inventory)
        financials['receivables'].append(receivables)
        financials['payables'].append(payables)
        financials['cash'].append(cash)
        financials['equity'].append(equity)
        financials['st_debt'].append(st_debt)
        financials['lt_debt'].append(lt_debt)
    
    return financials


# ================================
# SCREENER SPECIFIC UTILITIES
# ================================

def get_screener_shares_outstanding(df_bs, year_col):
    """
    Extract number of shares outstanding from Balance Sheet
    
    Returns:
        int: Number of shares
    """
    shares = get_value_from_screener_df(df_bs, 'No. of Equity Shares', year_col)
    if shares > 0:
        return int(shares)
    
    # Try without the period
    shares = get_value_from_screener_df(df_bs, 'No of Equity Shares', year_col)
    if shares > 0:
        return int(shares)
    
    return 0


def get_screener_face_value(df_bs, year_col):
    """
    Extract face value per share from Balance Sheet
    
    Returns:
        float: Face value
    """
    face_value = get_value_from_screener_df(df_bs, 'Face value', year_col)
    return face_value if face_value > 0 else 10.0  # Default to 10


# ================================
# DDM (DIVIDEND DISCOUNT MODEL) FOR SCREENER
# ================================

def calculate_screener_ddm_valuation(financials, num_shares, required_return=0.12, growth_rate=0.05):
    """
    Calculate DDM valuation using historical dividend data from Screener template
    
    Args:
        financials: Dict with 'dividends' key containing historical dividend amounts
        num_shares: Number of shares outstanding
        required_return: Required rate of return (default 12%)
        growth_rate: Expected dividend growth rate (default 5%)
    
    Returns:
        dict: DDM valuation results
    """
    if not financials or 'dividends' not in financials or num_shares <= 0:
        return None
    
    # Convert None/NaN to 0 for blank cells
    dividends = [float(d) if d and not (isinstance(d, float) and np.isnan(d)) else 0.0 for d in financials['dividends']]
    years = financials.get('years', [])
    
    # Filter out zero dividends for analysis
    non_zero_divs = [d for d in dividends if d > 0]
    
    if len(non_zero_divs) < 2:
        return {
            'model': 'DDM',
            'status': 'Insufficient Data',
            'message': 'Need at least 2 years of dividend history',
            'intrinsic_value_per_share': 0,
            'total_intrinsic_value': 0
        }
    
    # Calculate historical dividend growth rate (only for non-zero consecutive years)
    historical_growth_rates = []
    for i in range(1, len(dividends)):
        if dividends[i-1] > 0 and dividends[i] > 0:
            growth = (dividends[i] - dividends[i-1]) / dividends[i-1]
            historical_growth_rates.append(growth)
    
    avg_historical_growth = np.mean(historical_growth_rates) if historical_growth_rates else growth_rate
    
    # Use latest non-zero dividend
    latest_dividend = next((d for d in reversed(dividends) if d > 0), 0)
    
    if latest_dividend == 0:
        return {
            'model': 'DDM',
            'status': 'No Recent Dividend',
            'message': 'No dividend paid in recent years',
            'intrinsic_value_per_share': 0,
            'total_intrinsic_value': 0
        }
    
    dps = latest_dividend / num_shares  # Dividend per share
    
    # Gordon Growth Model: P = D1 / (r - g)
    # D1 = D0 * (1 + g)
    if required_return <= avg_historical_growth:
        return {
            'model': 'DDM (Gordon Growth Model)',
            'status': 'Invalid',
            'message': f'Required return ({required_return*100:.1f}%) must be greater than growth rate ({avg_historical_growth*100:.1f}%)',
            'intrinsic_value_per_share': 0,
            'total_intrinsic_value': 0,
            'latest_dps': dps,
            'historical_growth_rate': avg_historical_growth
        }
    
    # Calculate next year's expected dividend
    d1 = dps * (1 + avg_historical_growth)
    
    # Intrinsic value per share
    intrinsic_value_per_share = d1 / (required_return - avg_historical_growth)
    
    # Total intrinsic value
    total_intrinsic_value = intrinsic_value_per_share * num_shares
    
    return {
        'model': 'DDM (Gordon Growth Model)',
        'status': 'Success',
        'latest_dividend_total': latest_dividend,
        'latest_dps': dps,
        'num_shares': num_shares,
        'historical_growth_rate': avg_historical_growth,
        'assumed_growth_rate': avg_historical_growth,
        'required_return': required_return,
        'expected_next_dividend': d1,
        'intrinsic_value_per_share': intrinsic_value_per_share,
        'total_intrinsic_value': total_intrinsic_value,
        'dividend_history': dividends,
        'years': years
    }


# ================================
# RIM (RESIDUAL INCOME MODEL) FOR SCREENER
# ================================

def calculate_screener_rim_valuation(financials, num_shares, required_return=0.12, projection_years=5, terminal_growth=0.04):
    """
    Calculate RIM (Residual Income Model) valuation
    
    RIM Formula:
    Value = Book Value + PV(Expected Residual Income)
    Residual Income = Net Income - (Book Value * Required Return)
    
    Args:
        financials: Dict with historical financial data
        num_shares: Number of shares outstanding
        required_return: Required rate of return (default 12%)
        projection_years: Number of years to project (default 5)
        terminal_growth: Terminal growth rate (default 4%)
    
    Returns:
        dict: RIM valuation results
    """
    if not financials or num_shares <= 0:
        return None
    
    # Get latest book value and net income
    equity = financials['equity'][-1]  # Total equity (book value)
    net_income = financials['net_profit'][-1]  # Use reported net profit
    
    if equity <= 0:
        return {
            'model': 'RIM',
            'status': 'Invalid',
            'message': 'Book value must be positive',
            'intrinsic_value_per_share': 0,
            'total_intrinsic_value': 0
        }
    
    # Calculate historical ROE and growth rate
    if len(financials['equity']) >= 2 and len(financials['net_profit']) >= 2:
        roe_values = []
        for i in range(len(financials['equity'])):
            if financials['equity'][i] > 0:
                roe = financials['net_profit'][i] / financials['equity'][i]
                roe_values.append(roe)
        avg_roe = np.mean(roe_values) if roe_values else 0.15  # Default 15%
        
        # Calculate earnings growth rate
        earnings_growth_rates = []
        for i in range(1, len(financials['net_profit'])):
            if financials['net_profit'][i-1] > 0:
                growth = (financials['net_profit'][i] - financials['net_profit'][i-1]) / financials['net_profit'][i-1]
                earnings_growth_rates.append(growth)
        avg_earnings_growth = np.mean(earnings_growth_rates) if earnings_growth_rates else 0.08  # Default 8%
    else:
        avg_roe = 0.15
        avg_earnings_growth = 0.08
    
    # Project residual income
    residual_incomes = []
    projected_book_values = [equity]
    projected_earnings = [net_income]
    
    for year in range(1, projection_years + 1):
        # Project next year's book value and earnings
        next_book_value = projected_book_values[-1] * (1 + avg_earnings_growth)
        next_earnings = projected_earnings[-1] * (1 + avg_earnings_growth)
        
        # Calculate residual income
        # RI = Net Income - (Required Return × Book Value)
        required_income = projected_book_values[-1] * required_return
        residual_income = next_earnings - required_income
        
        residual_incomes.append(residual_income)
        projected_book_values.append(next_book_value)
        projected_earnings.append(next_earnings)
    
    # Calculate present value of residual incomes
    pv_residual_incomes = []
    for i, ri in enumerate(residual_incomes, 1):
        pv = ri / ((1 + required_return) ** i)
        pv_residual_incomes.append(pv)
    
    # Terminal value (perpetuity)
    if required_return > terminal_growth:
        terminal_ri = residual_incomes[-1] * (1 + terminal_growth)
        terminal_value = terminal_ri / (required_return - terminal_growth)
        pv_terminal = terminal_value / ((1 + required_return) ** projection_years)
    else:
        pv_terminal = 0
    
    # Total intrinsic value
    # Value = Current Book Value + PV(Residual Incomes) + PV(Terminal Value)
    total_intrinsic_value = equity + sum(pv_residual_incomes) + pv_terminal
    
    # Per share value
    intrinsic_value_per_share = total_intrinsic_value / num_shares
    book_value_per_share = equity / num_shares
    
    return {
        'model': 'RIM (Residual Income Model)',
        'status': 'Success',
        'current_book_value': equity,
        'book_value_per_share': book_value_per_share,
        'current_net_income': net_income,
        'num_shares': num_shares,
        'avg_roe': avg_roe,
        'avg_earnings_growth': avg_earnings_growth,
        'required_return': required_return,
        'projection_years': projection_years,
        'terminal_growth': terminal_growth,
        'residual_incomes': residual_incomes,
        'pv_residual_incomes': pv_residual_incomes,
        'pv_terminal_value': pv_terminal,
        'intrinsic_value_per_share': intrinsic_value_per_share,
        'total_intrinsic_value': total_intrinsic_value,
        'projected_book_values': projected_book_values,
        'projected_earnings': projected_earnings
    }


# ================================
# SCREENER EXCEL DOWNLOAD GENERATOR
# ================================

def generate_screener_valuation_excel(company_name, financials, dcf_results, ddm_results=None, rim_results=None, 
                                      comp_val_results=None, peer_comparison=None):
    """
    Generate comprehensive Excel report for Screener mode valuations
    
    Creates an Excel workbook with multiple sheets:
    1. Summary - Key valuation metrics
    2. Historical Financials - 3 years of historical data
    3. DCF Valuation - Detailed DCF model
    4. DDM Valuation - Dividend Discount Model (if applicable)
    5. RIM Valuation - Residual Income Model
    6. Comparative Valuation - Peer multiples
    7. Peer Comparison - Peer analysis dashboard
    
    Returns:
        BytesIO: Excel file buffer
    """
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== SHEET 1: SUMMARY =====
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = f"{company_name} - Valuation Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)
    
    row = 3
    ws_summary[f'A{row}'] = "Valuation Method"
    ws_summary[f'B{row}'] = "Intrinsic Value per Share"
    ws_summary[f'C{row}'] = "Total Intrinsic Value"
    for col in ['A', 'B', 'C']:
        ws_summary[f'{col}{row}'].fill = header_fill
        ws_summary[f'{col}{row}'].font = header_font
    
    row += 1
    
    # DCF
    if dcf_results:
        ws_summary[f'A{row}'] = "DCF (FCFF)"
        ws_summary[f'B{row}'] = dcf_results.get('intrinsic_value_per_share', 0)
        ws_summary[f'C{row}'] = dcf_results.get('enterprise_value', 0)
        row += 1
    
    # DDM
    if ddm_results and ddm_results.get('status') == 'Success':
        ws_summary[f'A{row}'] = "DDM (Dividend Discount)"
        ws_summary[f'B{row}'] = ddm_results.get('intrinsic_value_per_share', 0)
        ws_summary[f'C{row}'] = ddm_results.get('total_intrinsic_value', 0)
        row += 1
    
    # RIM
    if rim_results and rim_results.get('status') == 'Success':
        ws_summary[f'A{row}'] = "RIM (Residual Income)"
        ws_summary[f'B{row}'] = rim_results.get('intrinsic_value_per_share', 0)
        ws_summary[f'C{row}'] = rim_results.get('total_intrinsic_value', 0)
        row += 1
    
    # Comparative Valuation
    if comp_val_results and comp_val_results.get('valuations'):
        ws_summary[f'A{row}'] = "Peer Multiples (Avg)"
        avg_val = np.mean([v for v in comp_val_results['valuations'].values() if v > 0])
        ws_summary[f'B{row}'] = avg_val
        ws_summary[f'C{row}'] = avg_val * financials.get('num_shares', 1)
        row += 1
    
    # Column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 25
    
    # ===== SHEET 2: HISTORICAL FINANCIALS =====
    ws_hist = wb.create_sheet("Historical Financials", 1)
    ws_hist['A1'] = "Historical Financial Data (in Crores)"
    ws_hist['A1'].font = Font(bold=True, size=12)
    
    row = 3
    # Headers
    ws_hist['A3'] = "Metric"
    years = financials.get('years', [])
    for i, year in enumerate(years, 1):
        col_letter = chr(65 + i)  # B, C, D...
        ws_hist[f'{col_letter}3'] = year.replace('_', '')
        ws_hist[f'{col_letter}3'].fill = header_fill
        ws_hist[f'{col_letter}3'].font = header_font
    
    # P&L Section
    row = 4
    pl_metrics = [
        ('Revenue', 'revenue'),
        ('COGS', 'cogs'),
        ('Operating Expenses', 'opex'),
        ('EBITDA', 'ebitda'),
        ('Depreciation', 'depreciation'),
        ('EBIT', 'ebit'),
        ('Interest', 'interest'),
        ('Tax', 'tax'),
        ('Net Profit', 'net_profit'),
        ('NOPAT', 'nopat'),
        ('Dividends', 'dividends'),
    ]
    
    for metric_name, metric_key in pl_metrics:
        ws_hist[f'A{row}'] = metric_name
        for i, year in enumerate(years, 1):
            col_letter = chr(65 + i)
            value = financials[metric_key][i-1] if i <= len(financials[metric_key]) else 0
            ws_hist[f'{col_letter}{row}'] = value
        row += 1
    
    row += 1
    ws_hist[f'A{row}'] = "BALANCE SHEET"
    ws_hist[f'A{row}'].font = Font(bold=True)
    row += 1
    
    bs_metrics = [
        ('Fixed Assets', 'fixed_assets'),
        ('Inventory', 'inventory'),
        ('Receivables', 'receivables'),
        ('Payables', 'payables'),
        ('Cash', 'cash'),
        ('Equity', 'equity'),
        ('Short-term Debt', 'st_debt'),
        ('Long-term Debt', 'lt_debt'),
    ]
    
    for metric_name, metric_key in bs_metrics:
        ws_hist[f'A{row}'] = metric_name
        for i, year in enumerate(years, 1):
            col_letter = chr(65 + i)
            value = financials[metric_key][i-1] if i <= len(financials[metric_key]) else 0
            ws_hist[f'{col_letter}{row}'] = value
        row += 1
    
    ws_hist.column_dimensions['A'].width = 25
    
    # ===== SHEET 3: DCF VALUATION =====
    if dcf_results:
        ws_dcf = wb.create_sheet("DCF Valuation", 2)
        ws_dcf['A1'] = "DCF Valuation Model"
        ws_dcf['A1'].font = Font(bold=True, size=12)
        
        row = 3
        ws_dcf[f'A{row}'] = "WACC Calculation"
        ws_dcf[f'A{row}'].font = Font(bold=True)
        row += 1
        
        wacc_items = [
            ('Cost of Equity', dcf_results.get('cost_of_equity', 0)),
            ('Cost of Debt', dcf_results.get('cost_of_debt', 0)),
            ('Tax Rate', dcf_results.get('tax_rate', 0)),
            ('Weight of Equity', dcf_results.get('weight_equity', 0)),
            ('Weight of Debt', dcf_results.get('weight_debt', 0)),
            ('WACC', dcf_results.get('wacc', 0)),
        ]
        
        for item_name, item_value in wacc_items:
            ws_dcf[f'A{row}'] = item_name
            ws_dcf[f'B{row}'] = item_value
            row += 1
        
        row += 1
        ws_dcf[f'A{row}'] = "Valuation Results"
        ws_dcf[f'A{row}'].font = Font(bold=True)
        row += 1
        
        val_items = [
            ('Enterprise Value', dcf_results.get('enterprise_value', 0)),
            ('Less: Net Debt', dcf_results.get('net_debt', 0)),
            ('Equity Value', dcf_results.get('equity_value', 0)),
            ('Number of Shares', dcf_results.get('num_shares', 0)),
            ('Intrinsic Value per Share', dcf_results.get('intrinsic_value_per_share', 0)),
        ]
        
        for item_name, item_value in val_items:
            ws_dcf[f'A{row}'] = item_name
            ws_dcf[f'B{row}'] = item_value
            row += 1
        
        ws_dcf.column_dimensions['A'].width = 30
        ws_dcf.column_dimensions['B'].width = 20
    
    # ===== SHEET 4: DDM VALUATION =====
    if ddm_results and ddm_results.get('status') == 'Success':
        ws_ddm = wb.create_sheet("DDM Valuation", 3)
        ws_ddm['A1'] = "Dividend Discount Model (DDM)"
        ws_ddm['A1'].font = Font(bold=True, size=12)
        
        row = 3
        ddm_items = [
            ('Latest Dividend per Share', ddm_results.get('latest_dps', 0)),
            ('Historical Growth Rate', ddm_results.get('historical_growth_rate', 0)),
            ('Required Return', ddm_results.get('required_return', 0)),
            ('Expected Next Dividend', ddm_results.get('expected_next_dividend', 0)),
            ('Intrinsic Value per Share', ddm_results.get('intrinsic_value_per_share', 0)),
            ('Total Intrinsic Value', ddm_results.get('total_intrinsic_value', 0)),
        ]
        
        for item_name, item_value in ddm_items:
            ws_ddm[f'A{row}'] = item_name
            ws_ddm[f'B{row}'] = item_value
            row += 1
        
        # Dividend history
        row += 1
        ws_ddm[f'A{row}'] = "Dividend History"
        ws_ddm[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws_ddm[f'A{row}'] = "Year"
        ws_ddm[f'B{row}'] = "Dividend Amount"
        row += 1
        
        div_years = ddm_results.get('years', [])
        div_amounts = ddm_results.get('dividend_history', [])
        for i, (year, amount) in enumerate(zip(div_years, div_amounts)):
            ws_ddm[f'A{row}'] = year.replace('_', '')
            ws_ddm[f'B{row}'] = amount
            row += 1
        
        ws_ddm.column_dimensions['A'].width = 30
        ws_ddm.column_dimensions['B'].width = 20
    
    # ===== SHEET 5: RIM VALUATION =====
    if rim_results and rim_results.get('status') == 'Success':
        ws_rim = wb.create_sheet("RIM Valuation", 4)
        ws_rim['A1'] = "Residual Income Model (RIM)"
        ws_rim['A1'].font = Font(bold=True, size=12)
        
        row = 3
        rim_items = [
            ('Current Book Value', rim_results.get('current_book_value', 0)),
            ('Book Value per Share', rim_results.get('book_value_per_share', 0)),
            ('Current Net Income', rim_results.get('current_net_income', 0)),
            ('Average ROE', rim_results.get('avg_roe', 0)),
            ('Average Earnings Growth', rim_results.get('avg_earnings_growth', 0)),
            ('Required Return', rim_results.get('required_return', 0)),
            ('PV of Terminal Value', rim_results.get('pv_terminal_value', 0)),
            ('Intrinsic Value per Share', rim_results.get('intrinsic_value_per_share', 0)),
            ('Total Intrinsic Value', rim_results.get('total_intrinsic_value', 0)),
        ]
        
        for item_name, item_value in rim_items:
            ws_rim[f'A{row}'] = item_name
            ws_rim[f'B{row}'] = item_value
            row += 1
        
        ws_rim.column_dimensions['A'].width = 30
        ws_rim.column_dimensions['B'].width = 20
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


# ================================
# SCREENER MODE DISPLAY FUNCTIONS
# ================================

def display_screener_financial_summary(financials):
    """Display financial summary for Screener mode"""
    st.subheader("📊 Financial Summary (Last 3 Years)")
    
    years = financials.get('years', [])
    
    # Create DataFrame for display
    summary_data = {
        'Metric': [
            'Revenue', 'COGS', 'Operating Expenses', 'EBITDA', 'EBIT',
            'Net Profit', 'Dividends', 'Fixed Assets', 'Inventory',
            'Receivables', 'Cash', 'Total Equity', 'Total Debt'
        ],
        years[0].replace('_', ''): [
            financials['revenue'][0], financials['cogs'][0], financials['opex'][0],
            financials['ebitda'][0], financials['ebit'][0], financials['net_profit'][0],
            financials['dividends'][0], financials['fixed_assets'][0], financials['inventory'][0],
            financials['receivables'][0], financials['cash'][0], financials['equity'][0],
            financials['st_debt'][0] + financials['lt_debt'][0]
        ] if len(years) > 0 else [],
        years[1].replace('_', ''): [
            financials['revenue'][1], financials['cogs'][1], financials['opex'][1],
            financials['ebitda'][1], financials['ebit'][1], financials['net_profit'][1],
            financials['dividends'][1], financials['fixed_assets'][1], financials['inventory'][1],
            financials['receivables'][1], financials['cash'][1], financials['equity'][1],
            financials['st_debt'][1] + financials['lt_debt'][1]
        ] if len(years) > 1 else [],
        years[2].replace('_', ''): [
            financials['revenue'][2], financials['cogs'][2], financials['opex'][2],
            financials['ebitda'][2], financials['ebit'][2], financials['net_profit'][2],
            financials['dividends'][2], financials['fixed_assets'][2], financials['inventory'][2],
            financials['receivables'][2], financials['cash'][2], financials['equity'][2],
            financials['st_debt'][2] + financials['lt_debt'][2]
        ] if len(years) > 2 else []
    }
    
    df_summary = pd.DataFrame(summary_data)
    st.dataframe(df_summary, use_container_width=True)


def display_screener_ddm_results(ddm_results):
    """Display DDM valuation results"""
    if not ddm_results:
        return
    
    st.subheader("💰 Dividend Discount Model (DDM) Valuation")
    
    if ddm_results.get('status') != 'Success':
        st.warning(f"⚠️ {ddm_results.get('message', 'Unable to calculate DDM')}")
        return
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Latest DPS", f"₹{ddm_results['latest_dps']:.2f}")
        st.metric("Historical Growth", f"{ddm_results['historical_growth_rate']*100:.1f}%")
    
    with col2:
        st.metric("Required Return", f"{ddm_results['required_return']*100:.1f}%")
        st.metric("Expected Next Dividend", f"₹{ddm_results['expected_next_dividend']:.2f}")
    
    with col3:
        st.metric("Intrinsic Value/Share", f"₹{ddm_results['intrinsic_value_per_share']:.2f}")
        st.metric("Total Value", f"₹{ddm_results['total_intrinsic_value']/1e7:.2f} Cr")
    
    # Dividend history chart
    if ddm_results.get('dividend_history') and ddm_results.get('years'):
        import plotly.graph_objects as go
        
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=[y.replace('_', '') for y in ddm_results['years']],
            y=ddm_results['dividend_history'],
            name='Dividend Amount',
            marker_color='#1f77b4'
        ))
        fig.update_layout(
            title="Historical Dividend Trend",
            xaxis_title="Year",
            yaxis_title="Dividend Amount (₹ Cr)",
            height=400
        )
        st.plotly_chart(fig, use_container_width=True)


def display_screener_rim_results(rim_results):
    """Display RIM valuation results"""
    if not rim_results:
        return
    
    st.subheader("📈 Residual Income Model (RIM) Valuation")
    
    if rim_results.get('status') != 'Success':
        st.warning(f"⚠️ {rim_results.get('message', 'Unable to calculate RIM')}")
        return
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Book Value/Share", f"₹{rim_results['book_value_per_share']:.2f}")
        st.metric("Average ROE", f"{rim_results['avg_roe']*100:.1f}%")
    
    with col2:
        st.metric("Avg Earnings Growth", f"{rim_results['avg_earnings_growth']*100:.1f}%")
        st.metric("Required Return", f"{rim_results['required_return']*100:.1f}%")
    
    with col3:
        st.metric("Intrinsic Value/Share", f"₹{rim_results['intrinsic_value_per_share']:.2f}")
        st.metric("Total Value", f"₹{rim_results['total_intrinsic_value']/1e7:.2f} Cr")
    
    # Show residual income breakdown
    if rim_results.get('residual_incomes') and rim_results.get('pv_residual_incomes'):
        st.markdown("#### Projected Residual Income")
        
        ri_data = {
            'Year': list(range(1, len(rim_results['residual_incomes']) + 1)),
            'Residual Income': rim_results['residual_incomes'],
            'Present Value': rim_results['pv_residual_incomes']
        }
        df_ri = pd.DataFrame(ri_data)
        st.dataframe(df_ri, use_container_width=True)


# ================================
# EXPORT FUNCTIONS
# ================================

# All export and display functions will be imported from main file
# This module focuses on Screener-specific parsing and calculations
